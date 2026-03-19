#!/usr/bin/env python3
"""
ETL Script: Migration Excel → Supabase/PostgreSQL
Projet: Glyc Pro (ex-Prediabaid)

Lit les fichiers Excel de prediabaid-decision-server/data/menus/
et les insère dans la base Supabase.

Usage:
  # Dry-run (génère le SQL sans l'exécuter)
  python3 migrate_excel_to_supabase.py --dry-run

  # Exécution réelle
  python3 migrate_excel_to_supabase.py --db-url "postgresql://user:pass@host:port/dbname"

  # Via variables d'environnement
  export DATABASE_URL="postgresql://user:pass@db.xxx.supabase.co:5432/postgres"
  python3 migrate_excel_to_supabase.py

Prérequis:
  pip install openpyxl psycopg2-binary
"""

import os
import sys
import json
import logging
import argparse
from pathlib import Path
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "prediabaid-decision-server" / "data"
MENUS_DIR = DATA_DIR / "menus"
MESSAGES_DIR = DATA_DIR / "messages" / "menus"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("migrate")


# ---------------------------------------------------------------------------
# EXCEL READERS
# ---------------------------------------------------------------------------

def read_xlsx(path, sheet_name=None):
    """Read an xlsx file and return list of dicts (one per row)."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    target_sheet = sheet_name or wb.sheetnames[0]
    ws = wb[target_sheet]

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i] is not None:
                row_dict[headers[i]] = val
        rows.append(row_dict)
    return rows


def read_translations(filename):
    """Read a Messages_*.xlsx and return {id: {fr, en, es}} dict."""
    path = MESSAGES_DIR / filename
    if not path.exists():
        log.warning(f"Translation file not found: {path}")
        return {}

    rows = read_xlsx(path, "Message")
    translations = {}
    for row in rows:
        key = row.get("ID")
        if key is None:
            continue
        translations[str(key)] = {
            "fr": row.get("FR"),
            "en": row.get("EN"),
            "es": row.get("ES"),
            "l4": row.get("L4"),
            "l5": row.get("L5"),
            "l6": row.get("L6"),
        }
    return translations


# ---------------------------------------------------------------------------
# DATA TRANSFORMERS
# ---------------------------------------------------------------------------

def yn_to_bool(val):
    """Convert 'Y'/'N'/None to boolean."""
    if val is None:
        return False
    return str(val).strip().upper() == "Y"


def safe_int(val):
    """Convert to int or None."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def safe_float(val):
    """Convert to float or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Convert to stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def split_slash_ids(val):
    """Split slash-separated IDs like '01/03/05' into list."""
    if val is None:
        return []
    return [x.strip() for x in str(val).strip().rstrip("/").split("/") if x.strip()]


def parse_type_repas(val):
    """Parse TypeRepasPredef like 'DEJEUNER ET DINER' into PG array literal."""
    if val is None:
        return "{}"
    mapping = {
        "PETIT_DEJEUNER": "PETIT_DEJEUNER",
        "SNACK_AM": "SNACK_AM",
        "SNACK_PM": "SNACK_PM",
        "DEJEUNER": "DEJEUNER",
        "DINER": "DINER",
    }
    parts = str(val).replace(" ET ", ",").replace(" et ", ",").split(",")
    result = []
    for p in parts:
        p = p.strip()
        if p in mapping:
            result.append(mapping[p])
    return "{" + ",".join(result) + "}"


# ---------------------------------------------------------------------------
# EXTRACT & TRANSFORM
# ---------------------------------------------------------------------------

def extract_aliment_categories(translations):
    """Extract from Categories_Aliments.xlsx."""
    rows = read_xlsx(MENUS_DIR / "Categories_Aliments.xlsx", "CATEGORIES ALIMENTS")
    categories = []
    for row in rows:
        cat_id = safe_str(row.get("id"))
        if cat_id is None:
            continue
        name_key = safe_str(row.get("name")) or ""
        tr = translations.get(name_key, {})
        categories.append({
            "id": cat_id,
            "name_key": name_key,
            "categorie_generale": safe_str(row.get("CategorieGenerale")),
            "rubrique": safe_str(row.get("rubrique")),
            "min_hebdo": safe_int(row.get("minHebdo")),
            "max_hebdo": safe_int(row.get("maxHebdo")),
            "est_casher": yn_to_bool(row.get("estCasher")),
            "est_vegetarien": yn_to_bool(row.get("estVegetarien")),
            "est_halal": yn_to_bool(row.get("estHalal")),
            "sans_gluten": yn_to_bool(row.get("sansGluten")),
            "est_adaptable": yn_to_bool(row.get("estAdaptable")),
            "name_fr": tr.get("fr"),
            "name_en": tr.get("en"),
            "name_es": tr.get("es"),
        })
    log.info(f"Extracted {len(categories)} aliment categories")
    return categories


def extract_aliments(translations):
    """Extract from Aliments.xlsx."""
    rows = read_xlsx(MENUS_DIR / "Aliments.xlsx", "ALIMENTS")
    aliments = []
    for row in rows:
        aliment_id = safe_str(row.get("Id"))
        if aliment_id is None:
            continue
        tr = translations.get(aliment_id, {})
        aliments.append({
            "id": aliment_id,
            "category_id": safe_str(row.get("Category Id")),
            "food_class": safe_str(row.get("FoodClass")),
            "occurrence_class": safe_str(row.get("OccurrenceClass")),
            "season_start_month": safe_int(row.get("SeasonStartMonth")),
            "season_end_month": safe_int(row.get("SeasonEndMonth")),
            "energy_kcal": safe_float(row.get("Energy_Kcal")),
            "protein_g": safe_float(row.get("Protein_(g)")),
            "lipid_tot_g": safe_float(row.get("Lipid_Tot_(g)")),
            "carbohydrt_g": safe_float(row.get("Carbohydrt_(g)")),
            "fiber_td_g": safe_float(row.get("Fiber_TD_(g)")),
            "sugar_tot_g": safe_float(row.get("Sugar_Tot_(g)")),
            "fa_sat_g": safe_float(row.get("FA_Sat_(g)")),
            "fa_mono_g": safe_float(row.get("FA_Mono_(g)")),
            "fa_poly_g": safe_float(row.get("FA_Poly_(g)")),
            "name_fr": tr.get("fr"),
            "name_en": tr.get("en"),
            "name_es": tr.get("es"),
        })
    log.info(f"Extracted {len(aliments)} aliments")
    return aliments


def extract_recipes(translations):
    """Extract from Recettes.xlsx → recipes, ingredients, phase conditions."""
    rows = read_xlsx(MENUS_DIR / "Recettes.xlsx", "Recettes")

    recipes = {}          # id → recipe dict
    ingredients = []      # list of ingredient dicts
    conditions = []       # list of condition dicts

    ingredient_order = defaultdict(int)  # recipe_id → sort counter

    for row in rows:
        recipe_id = safe_str(row.get("id"))
        if recipe_id is None:
            continue

        # New recipe?
        if recipe_id not in recipes:
            tr = translations.get(recipe_id, {})
            recipes[recipe_id] = {
                "id": recipe_id,
                "description_key": safe_str(row.get("description")),
                "id_photo": safe_str(row.get("idPhoto")),
                "duree_preparation": safe_int(row.get("DureePreparation")),
                "duree_cuisson": safe_int(row.get("DureeCuisson")),
                "nb_kcal_journalier_definition": safe_int(row.get("nbKCalJournalierDefinition")),
                "composant_repas": safe_str(row.get("ComposantRepasPredef")),
                "fete": safe_str(row.get("Fete")),
                "name_fr": tr.get("fr"),
                "name_en": tr.get("en"),
                "name_es": tr.get("es"),
            }

        # Ingredient?
        aliment_id = safe_str(row.get("idAliment"))
        if aliment_id is not None:
            ingredient_order[recipe_id] += 1
            ingredients.append({
                "recipe_id": recipe_id,
                "aliment_id": aliment_id,
                "portion": safe_float(row.get("Portion")) or 0,
                "unite": safe_str(row.get("Unite")) or "unit",
                "sort_order": ingredient_order[recipe_id],
            })

        # Phase/kcal condition?
        type_repas_predef = safe_str(row.get("TypeRepasPredef"))
        nb_kcal_activation = safe_int(row.get("nbKCalJournalierActivation"))
        nb_kcal_max = safe_int(row.get("nbKCalJournalierMaximun"))
        phase_activation = safe_str(row.get("phaseActivation"))

        has_condition = any([type_repas_predef, nb_kcal_activation, nb_kcal_max, phase_activation])
        if has_condition:
            conditions.append({
                "recipe_id": recipe_id,
                "types_repas": parse_type_repas(type_repas_predef),
                "nb_kcal_activation": nb_kcal_activation,
                "nb_kcal_maximum": nb_kcal_max,
                "phase_activation": phase_activation,
            })

    log.info(f"Extracted {len(recipes)} recipes, {len(ingredients)} ingredients, {len(conditions)} phase conditions")
    return list(recipes.values()), ingredients, conditions


def extract_allergies():
    """Extract from Allergies.xlsx."""
    rows = read_xlsx(MENUS_DIR / "Allergies.xlsx", "Allergies")
    allergies = []
    excluded_categories = []
    excluded_aliments = []

    for row in rows:
        allergy_id = safe_str(row.get("Id"))
        if allergy_id is None:
            continue

        allergies.append({
            "id": allergy_id,
            "reponse_bilan": safe_str(row.get("ReponseBilan")) or "",
        })

        for cat_id in split_slash_ids(row.get("Categories")):
            excluded_categories.append({
                "allergy_id": allergy_id,
                "category_id": cat_id,
            })

        for alim_id in split_slash_ids(row.get("Aliments")):
            excluded_aliments.append({
                "allergy_id": allergy_id,
                "aliment_id": alim_id,
            })

    log.info(f"Extracted {len(allergies)} allergies, {len(excluded_categories)} excluded categories, {len(excluded_aliments)} excluded aliments")
    return allergies, excluded_categories, excluded_aliments


def extract_kosher_incompatibilities():
    """Extract from Aliments_Incompatibles_Casher.xlsx."""
    rows = read_xlsx(MENUS_DIR / "Aliments_Incompatibles_Casher.xlsx", "Aliments incompatibles")
    incompatibilities = []
    for row in rows:
        cat1 = safe_str(row.get("CAT ID 1"))
        cat2 = safe_str(row.get("CAT ID 2"))
        if cat1 and cat2:
            incompatibilities.append({
                "category_id_1": cat1,
                "category_id_2": cat2,
            })
    log.info(f"Extracted {len(incompatibilities)} kosher incompatibilities")
    return incompatibilities


def extract_phase_restrictions():
    """Extract from RestrictionPhaseAlimentCategory.xlsx."""
    rows = read_xlsx(MENUS_DIR / "RestrictionPhaseAlimentCategory.xlsx", "RestrictionPhaseAlimentCategory")
    restrictions = []
    for row in rows:
        phase_id = safe_str(row.get("idRestrictedPhase"))
        cat_id = safe_str(row.get("idRestrictedCategory"))
        if phase_id and cat_id:
            restrictions.append({
                "phase_id": phase_id,
                "category_id": str(cat_id),
            })
    log.info(f"Extracted {len(restrictions)} phase restrictions")
    return restrictions


def extract_unit_conversions(translations):
    """Extract from UnitConvert.xlsx."""
    rows = read_xlsx(MENUS_DIR / "UnitConvert.xlsx", "UnitConvert")
    conversions = []
    for row in rows:
        unit = safe_str(row.get("Unit"))
        if unit is None:
            continue
        unit_key = f"unit.{unit}"
        tr = translations.get(unit_key, {})
        conversions.append({
            "unit": unit,
            "category_name": safe_str(row.get("Category")),
            "category_id": safe_str(row.get("IdCategory")),
            "aliment_name": safe_str(row.get("Aliment")),
            "aliment_id": safe_str(row.get("IdAliment")),
            "priority": safe_float(row.get("Priority")) or 0,
            "systems": safe_str(row.get("Systems")) or "ALL",
            "grams": safe_float(row.get("g")),
            "ml": safe_float(row.get("ml")),
            "round_precision": safe_float(row.get("Round")),
            "discrete_fractionals": safe_str(row.get("DiscreteFractionals")),
            "name_fr": tr.get("fr"),
            "name_en": tr.get("en"),
        })
    log.info(f"Extracted {len(conversions)} unit conversions")
    return conversions


def extract_unit_system_mappings():
    """Extract from UnitConvert.xlsx, sheet UnitSystemResponseMapping."""
    rows = read_xlsx(MENUS_DIR / "UnitConvert.xlsx", "UnitSystemResponseMapping")
    mappings = []
    for row in rows:
        key = safe_str(row.get("Response"))
        system = safe_str(row.get("UnitSystemAbrev"))
        if key and system:
            mappings.append({
                "response_key": key,
                "unit_system": system,
            })
    log.info(f"Extracted {len(mappings)} unit system mappings")
    return mappings


def extract_menu_definitions():
    """Extract from Menus.xlsx (multi-sheet)."""
    wb = openpyxl.load_workbook(str(MENUS_DIR / "Menus.xlsx"), data_only=True)

    phase_map = {
        "MENUS DETOX": "DETOX",
        "MENUS INTENSIVE": "INTENSIVE",
        "MENUS PVBF": "PVBF",
        "MENUS PROGRESSIVE": "PROGRESSIVE",
        "MENUS STABI": "STABILISATION",
    }

    type_repas_order = ["PETIT_DEJEUNER", "SNACK_AM", "DEJEUNER", "SNACK_PM", "DINER"]

    definitions = []

    for sheet_name, phase_id in phase_map.items():
        ws = wb[sheet_name]

        # Read phase kcal ratio (cell B1, if present)
        phase_ratio = safe_float(ws.cell(1, 2).value)

        # Parse blocks: each meal type block starts with the meal type name
        current_type_repas_idx = -1
        kcal_headers = []

        for row_idx in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            first_val = safe_str(row_vals[0]) if row_vals else None

            if first_val is None:
                continue

            # Check if this row is a meal type marker
            if first_val in type_repas_order:
                current_type_repas_idx = type_repas_order.index(first_val)
                continue

            # Check if this is a RUBRIQUES header row
            if first_val == "RUBRIQUES":
                kcal_headers = []
                for c in range(2, len(row_vals)):
                    val = safe_str(row_vals[c])
                    if val and val.startswith("Calories /"):
                        try:
                            kcal_headers.append(int(val.replace("Calories /", "")))
                        except ValueError:
                            pass
                continue

            # Skip TOTAL rows
            if first_val.startswith("TOTAL"):
                continue

            # This should be a rubrique data row
            if current_type_repas_idx < 0 or not kcal_headers:
                continue

            rubrique_label = first_val
            is_active = yn_to_bool(row_vals[1]) if len(row_vals) > 1 else False

            # Build kcal targets dict
            kcal_targets = {}
            for i, kcal_level in enumerate(kcal_headers):
                col_idx = i + 2  # offset: col 0=rubrique, col 1=O/N, col 2+=kcal values
                if col_idx < len(row_vals):
                    val = row_vals[col_idx]
                    if val is not None and val != "N/A":
                        kcal_val = safe_int(val)
                        if kcal_val is not None:
                            kcal_targets[str(kcal_level)] = kcal_val

            definitions.append({
                "phase": phase_id,
                "phase_kcal_ratio": phase_ratio,
                "type_repas": type_repas_order[current_type_repas_idx],
                "rubrique_label": rubrique_label,
                "is_active": is_active,
                "kcal_targets": json.dumps(kcal_targets),
            })

    log.info(f"Extracted {len(definitions)} menu definitions")
    return definitions


def extract_category_correspondences():
    """Extract from Correspondances_Categories_Aliments.xlsx."""
    rows = read_xlsx(MENUS_DIR / "Correspondances_Categories_Aliments.xlsx", "CATEGORIES ALIMENTS")
    correspondences = []
    current_group = None
    for row in rows:
        name = safe_str(row.get("NAME"))
        if name:
            current_group = name
        cat_id = safe_str(row.get("id"))
        if cat_id is None:
            continue
        correspondences.append({
            "category_id": cat_id,
            "group_name": current_group,
            "supports_diet_preferences": yn_to_bool(row.get("Diet Preferences ")),
            "supports_ingredients": yn_to_bool(row.get("Ingredients")),
        })
    log.info(f"Extracted {len(correspondences)} category correspondences")
    return correspondences


def extract_generic_translations():
    """Extract translations that don't map to specific entities."""
    all_translations = {}

    # Collect all message files
    for fname in ["Messages_Recettes.xlsx", "Messages_Aliments.xlsx",
                   "Messages_Categories.xlsx", "Messages_Unites.xlsx"]:
        path = MESSAGES_DIR / fname
        if path.exists():
            tr = read_translations(fname)
            all_translations.update(tr)

    result = []
    for key, vals in all_translations.items():
        result.append({
            "key": key,
            "fr": vals.get("fr"),
            "en": vals.get("en"),
            "es": vals.get("es"),
            "l4": vals.get("l4"),
            "l5": vals.get("l5"),
            "l6": vals.get("l6"),
        })
    log.info(f"Extracted {len(result)} total translations")
    return result


# ---------------------------------------------------------------------------
# SQL GENERATION
# ---------------------------------------------------------------------------

def sql_val(val):
    """Format a Python value as a SQL literal."""
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        return str(val)
    # String: escape single quotes
    s = str(val).replace("'", "''")
    return f"'{s}'"


def generate_insert(table, columns, rows):
    """Generate a bulk INSERT statement."""
    if not rows:
        return ""

    lines = [f"INSERT INTO {table} ({', '.join(columns)}) VALUES"]
    value_lines = []
    for row in rows:
        vals = ", ".join(sql_val(row.get(col)) for col in columns)
        value_lines.append(f"  ({vals})")
    lines.append(",\n".join(value_lines))
    lines.append("ON CONFLICT DO NOTHING;\n")
    return "\n".join(lines)


def generate_full_sql(data):
    """Generate the complete SQL migration script."""
    sql_parts = []

    sql_parts.append("-- =============================================================")
    sql_parts.append("-- MIGRATION ETL: Excel → PostgreSQL/Supabase")
    sql_parts.append("-- Generated by migrate_excel_to_supabase.py")
    sql_parts.append("-- =============================================================\n")
    sql_parts.append("BEGIN;\n")

    # Truncate in reverse dependency order
    sql_parts.append("-- Clear existing data (reverse FK order)")
    tables_to_clear = [
        "translations", "category_correspondences", "menu_definitions",
        "unit_system_mappings", "unit_conversions",
        "phase_category_restrictions", "kosher_incompatibilities",
        "allergy_excluded_aliments", "allergy_excluded_categories", "allergies",
        "recipe_phase_conditions", "recipe_ingredients", "recipes",
        "aliments", "aliment_categories",
    ]
    for t in tables_to_clear:
        sql_parts.append(f"TRUNCATE TABLE {t} CASCADE;")
    sql_parts.append("")

    # 1. Aliment categories
    sql_parts.append("-- 1. Aliment categories")
    sql_parts.append(generate_insert(
        "aliment_categories",
        ["id", "name_key", "categorie_generale", "rubrique", "min_hebdo", "max_hebdo",
         "est_casher", "est_vegetarien", "est_halal", "sans_gluten", "est_adaptable",
         "name_fr", "name_en", "name_es"],
        data["categories"],
    ))

    # 2. Aliments
    sql_parts.append("-- 2. Aliments")
    sql_parts.append(generate_insert(
        "aliments",
        ["id", "category_id", "food_class", "occurrence_class",
         "season_start_month", "season_end_month",
         "energy_kcal", "protein_g", "lipid_tot_g", "carbohydrt_g",
         "fiber_td_g", "sugar_tot_g", "fa_sat_g", "fa_mono_g", "fa_poly_g",
         "name_fr", "name_en", "name_es"],
        data["aliments"],
    ))

    # 3. Recipes
    sql_parts.append("-- 3. Recipes")
    sql_parts.append(generate_insert(
        "recipes",
        ["id", "description_key", "id_photo", "duree_preparation", "duree_cuisson",
         "nb_kcal_journalier_definition", "composant_repas", "fete",
         "name_fr", "name_en", "name_es"],
        data["recipes"],
    ))

    # 4. Recipe ingredients
    sql_parts.append("-- 4. Recipe ingredients")
    sql_parts.append(generate_insert(
        "recipe_ingredients",
        ["recipe_id", "aliment_id", "portion", "unite", "sort_order"],
        data["ingredients"],
    ))

    # 5. Recipe phase conditions
    sql_parts.append("-- 5. Recipe phase conditions")
    sql_parts.append(generate_insert(
        "recipe_phase_conditions",
        ["recipe_id", "types_repas", "nb_kcal_activation", "nb_kcal_maximum", "phase_activation"],
        data["conditions"],
    ))

    # 6. Allergies
    sql_parts.append("-- 6. Allergies")
    sql_parts.append(generate_insert(
        "allergies",
        ["id", "reponse_bilan"],
        data["allergies"],
    ))

    sql_parts.append("-- 6b. Allergy excluded categories")
    sql_parts.append(generate_insert(
        "allergy_excluded_categories",
        ["allergy_id", "category_id"],
        data["allergy_excluded_categories"],
    ))

    sql_parts.append("-- 6c. Allergy excluded aliments")
    sql_parts.append(generate_insert(
        "allergy_excluded_aliments",
        ["allergy_id", "aliment_id"],
        data["allergy_excluded_aliments"],
    ))

    # 7. Kosher incompatibilities
    sql_parts.append("-- 7. Kosher incompatibilities")
    sql_parts.append(generate_insert(
        "kosher_incompatibilities",
        ["category_id_1", "category_id_2"],
        data["kosher_incompatibilities"],
    ))

    # 8. Phase restrictions
    sql_parts.append("-- 8. Phase restrictions")
    sql_parts.append(generate_insert(
        "phase_category_restrictions",
        ["phase_id", "category_id"],
        data["phase_restrictions"],
    ))

    # 9. Unit conversions
    sql_parts.append("-- 9. Unit conversions")
    sql_parts.append(generate_insert(
        "unit_conversions",
        ["unit", "category_name", "category_id", "aliment_name", "aliment_id",
         "priority", "systems", "grams", "ml", "round_precision",
         "discrete_fractionals", "name_fr", "name_en"],
        data["unit_conversions"],
    ))

    # 10. Unit system mappings
    sql_parts.append("-- 10. Unit system mappings")
    sql_parts.append(generate_insert(
        "unit_system_mappings",
        ["response_key", "unit_system"],
        data["unit_system_mappings"],
    ))

    # 11. Menu definitions
    sql_parts.append("-- 11. Menu definitions")
    sql_parts.append(generate_insert(
        "menu_definitions",
        ["phase", "phase_kcal_ratio", "type_repas", "rubrique_label", "is_active", "kcal_targets"],
        data["menu_definitions"],
    ))

    # 12. Category correspondences
    sql_parts.append("-- 12. Category correspondences")
    sql_parts.append(generate_insert(
        "category_correspondences",
        ["category_id", "group_name", "supports_diet_preferences", "supports_ingredients"],
        data["category_correspondences"],
    ))

    # 13. Translations
    sql_parts.append("-- 13. Translations")
    sql_parts.append(generate_insert(
        "translations",
        ["key", "fr", "en", "es", "l4", "l5", "l6"],
        data["translations"],
    ))

    sql_parts.append("COMMIT;\n")

    return "\n".join(sql_parts)


# ---------------------------------------------------------------------------
# DATABASE EXECUTION
# ---------------------------------------------------------------------------

def execute_sql(db_url, sql):
    """Execute the SQL against a PostgreSQL database."""
    try:
        import psycopg2
    except ImportError:
        log.error("psycopg2 not installed. Run: pip install psycopg2-binary")
        sys.exit(1)

    log.info(f"Connecting to database...")
    conn = psycopg2.connect(db_url)
    try:
        cur = conn.cursor()
        cur.execute(sql)
        conn.commit()
        log.info("Migration completed successfully!")
    except Exception as e:
        conn.rollback()
        log.error(f"Migration failed, rolling back: {e}")
        raise
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# VALIDATION
# ---------------------------------------------------------------------------

def validate_data(data):
    """Run validation checks on extracted data."""
    errors = []
    warnings = []

    # Check category references in aliments
    cat_ids = {c["id"] for c in data["categories"]}
    for a in data["aliments"]:
        if a["category_id"] not in cat_ids:
            errors.append(f"Aliment {a['id']} references unknown category {a['category_id']}")

    # Check aliment references in ingredients
    aliment_ids = {a["id"] for a in data["aliments"]}
    for ing in data["ingredients"]:
        if ing["aliment_id"] not in aliment_ids:
            warnings.append(f"Ingredient in recipe {ing['recipe_id']} references unknown aliment {ing['aliment_id']}")

    # Check recipe references in ingredients
    recipe_ids = {r["id"] for r in data["recipes"]}
    for ing in data["ingredients"]:
        if ing["recipe_id"] not in recipe_ids:
            errors.append(f"Ingredient references unknown recipe {ing['recipe_id']}")

    # Check allergy category references
    for exc in data["allergy_excluded_categories"]:
        if exc["category_id"] not in cat_ids:
            warnings.append(f"Allergy {exc['allergy_id']} excludes unknown category {exc['category_id']}")

    # Report
    if errors:
        log.error(f"Validation found {len(errors)} errors:")
        for e in errors[:20]:
            log.error(f"  - {e}")

    if warnings:
        log.warning(f"Validation found {len(warnings)} warnings:")
        for w in warnings[:20]:
            log.warning(f"  - {w}")

    if not errors:
        log.info("Validation passed!")

    return len(errors) == 0


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Migrate Excel data to Supabase/PostgreSQL")
    parser.add_argument("--dry-run", action="store_true",
                        help="Generate SQL file without executing")
    parser.add_argument("--db-url", type=str, default=None,
                        help="PostgreSQL connection URL")
    parser.add_argument("--output", type=str, default="migration_data.sql",
                        help="Output SQL file path (default: migration_data.sql)")
    parser.add_argument("--skip-validation", action="store_true",
                        help="Skip data validation")
    args = parser.parse_args()

    db_url = args.db_url or os.environ.get("DATABASE_URL")

    if not args.dry_run and not db_url:
        log.error("Please provide --db-url or set DATABASE_URL environment variable")
        log.info("Or use --dry-run to generate SQL without executing")
        sys.exit(1)

    # Verify data files exist
    if not MENUS_DIR.exists():
        log.error(f"Data directory not found: {MENUS_DIR}")
        sys.exit(1)

    log.info("=" * 60)
    log.info("Starting ETL: Excel → PostgreSQL/Supabase")
    log.info("=" * 60)

    # ── EXTRACT ──────────────────────────────────────────────
    log.info("\n── EXTRACT ──")

    # Load translations first (needed for denormalization)
    tr_categories = read_translations("Messages_Categories.xlsx")
    tr_aliments = read_translations("Messages_Aliments.xlsx")
    tr_recettes = read_translations("Messages_Recettes.xlsx")
    tr_unites = read_translations("Messages_Unites.xlsx")

    data = {}
    data["categories"] = extract_aliment_categories(tr_categories)
    data["aliments"] = extract_aliments(tr_aliments)

    recipes, ingredients, conditions = extract_recipes(tr_recettes)
    data["recipes"] = recipes
    data["ingredients"] = ingredients
    data["conditions"] = conditions

    allergies, exc_cats, exc_alims = extract_allergies()
    data["allergies"] = allergies
    data["allergy_excluded_categories"] = exc_cats
    data["allergy_excluded_aliments"] = exc_alims

    data["kosher_incompatibilities"] = extract_kosher_incompatibilities()
    data["phase_restrictions"] = extract_phase_restrictions()
    data["unit_conversions"] = extract_unit_conversions(tr_unites)
    data["unit_system_mappings"] = extract_unit_system_mappings()
    data["menu_definitions"] = extract_menu_definitions()
    data["category_correspondences"] = extract_category_correspondences()
    data["translations"] = extract_generic_translations()

    # ── VALIDATE ─────────────────────────────────────────────
    if not args.skip_validation:
        log.info("\n── VALIDATE ──")
        is_valid = validate_data(data)
        if not is_valid:
            log.error("Validation failed. Fix errors above or use --skip-validation")
            sys.exit(1)

    # ── SUMMARY ──────────────────────────────────────────────
    log.info("\n── SUMMARY ──")
    log.info(f"  aliment_categories:          {len(data['categories']):>6}")
    log.info(f"  aliments:                    {len(data['aliments']):>6}")
    log.info(f"  recipes:                     {len(data['recipes']):>6}")
    log.info(f"  recipe_ingredients:          {len(data['ingredients']):>6}")
    log.info(f"  recipe_phase_conditions:     {len(data['conditions']):>6}")
    log.info(f"  allergies:                   {len(data['allergies']):>6}")
    log.info(f"  allergy_excluded_categories: {len(data['allergy_excluded_categories']):>6}")
    log.info(f"  allergy_excluded_aliments:   {len(data['allergy_excluded_aliments']):>6}")
    log.info(f"  kosher_incompatibilities:    {len(data['kosher_incompatibilities']):>6}")
    log.info(f"  phase_category_restrictions: {len(data['phase_restrictions']):>6}")
    log.info(f"  unit_conversions:            {len(data['unit_conversions']):>6}")
    log.info(f"  unit_system_mappings:        {len(data['unit_system_mappings']):>6}")
    log.info(f"  menu_definitions:            {len(data['menu_definitions']):>6}")
    log.info(f"  category_correspondences:    {len(data['category_correspondences']):>6}")
    log.info(f"  translations:                {len(data['translations']):>6}")

    # ── GENERATE SQL ─────────────────────────────────────────
    log.info("\n── GENERATE SQL ──")
    sql = generate_full_sql(data)

    # Always write the SQL file (useful for review)
    output_path = Path(args.output)
    output_path.write_text(sql, encoding="utf-8")
    log.info(f"SQL written to {output_path} ({len(sql):,} chars)")

    # ── EXECUTE ──────────────────────────────────────────────
    if args.dry_run:
        log.info("\nDry-run mode: SQL generated but not executed.")
        log.info(f"Review the file: {output_path}")
        log.info(f"Then run: python3 {__file__} --db-url \"postgresql://...\"")
    else:
        log.info("\n── EXECUTE ──")
        execute_sql(db_url, sql)

    log.info("\nDone!")


if __name__ == "__main__":
    main()
